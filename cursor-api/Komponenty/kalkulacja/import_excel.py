"""Import danych z pliku .xlsm (do analizy / GicleeArt3) do folderu data/."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from .store import data_dir

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]
    _IMPORT_ERR = exc
else:
    _IMPORT_ERR = None


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text in ("", "#N/A", "BRAK", " "):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def import_from_xlsm(source: str | Path) -> dict[str, int]:
    """Wczytuje arkusze z Excela i nadpisuje pliki JSON w data/."""
    if openpyxl is None:
        raise RuntimeError(
            "Brak biblioteki openpyxl — zainstaluj: pip install openpyxl"
        ) from _IMPORT_ERR

    path = Path(source)
    if not path.is_file():
        raise FileNotFoundError(f"Nie znaleziono pliku: {path}")

    out = data_dir()
    out.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(path, data_only=True)

    ws = wb["CENNIK MATERIAŁÓW"]
    materials: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1]:
            continue
        materials.append(
            {
                "id": str(row[1]).strip(),
                "category": row[2],
                "product": row[3],
                "size": row[4],
                "price": _to_float(row[5]) or 0,
            }
        )

    ws2 = wb["TABELA CEN WG MATERIAŁÓW"]
    helpers: dict[str, dict[str, float | None]] = {}
    for r in range(5, 11):
        key = ws2.cell(r, 9).value
        if key:
            helpers[str(key)] = {
                "A4": _to_float(ws2.cell(r, 10).value),
                "A3+": _to_float(ws2.cell(r, 11).value),
                "A2": _to_float(ws2.cell(r, 12).value),
            }

    price_table: list[dict[str, Any]] = []
    for row in ws2.iter_rows(min_row=13, values_only=True):
        if not row[1]:
            continue
        cost = _to_float(row[6])
        if cost is None:
            continue
        price_table.append(
            {
                "id_full": str(row[1]).strip(),
                "id_tr": str(row[2]).strip() if row[2] else "",
                "product": row[3],
                "format": row[4],
                "profile": row[5],
                "cost": cost,
            }
        )

    ws3 = wb["KALKULATOR KOSZTU DREWNA"]
    wood_defaults = {
        "price_per_meter": _to_float(ws3["D3"].value),
        "profile": ws3["D4"].value,
        "species": str(ws3["D5"].value or "").strip(),
        "format": ws3["D6"].value,
        "meters_per_frame": _to_float(ws3["D7"].value),
        "shipping": _to_float(ws3["E9"].value) or 25,
        "max_batch": 30,
        "max_gain_ratio": _to_float(ws3["H7"].value) or 0.2,
    }

    ws4 = wb["CENNIK"]
    existing_settings: dict[str, Any] = {}
    settings_path = out / "settings.json"
    if settings_path.is_file():
        try:
            existing_settings = json.loads(settings_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            existing_settings = {}
    settings = {
        "profile": ws4.cell(9, 4).value,
        "frame_type": ws4.cell(5, 5).value,
        "default_markup_pct": float(existing_settings.get("default_markup_pct") or 250),
        "default_production_minutes": float(existing_settings.get("default_production_minutes") or 45),
        "wood_origin": str(existing_settings.get("wood_origin") or "stolarz24"),
        "frames_per_day": float(existing_settings.get("frames_per_day") or 1.0),
        "frames_per_day_mode": str(existing_settings.get("frames_per_day_mode") or "manual"),
        "work_hours_per_day": float(existing_settings.get("work_hours_per_day") or 8.0),
        "work_days_per_month": float(existing_settings.get("work_days_per_month") or 22.0),
        "variant_pricing": dict(existing_settings.get("variant_pricing") or {}),
        "variant_production_minutes": dict(existing_settings.get("variant_production_minutes") or {}),
        "business_costs": dict(existing_settings.get("business_costs") or {}),
    }

    def row_costs(row_idx: int) -> dict[str, float]:
        result: dict[str, float] = {}
        for key, col in [
            ("A4_sosna", 5),
            ("A3+_sosna", 6),
            ("A2_sosna", 7),
            ("A4_dab", 8),
            ("A3+_dab", 9),
            ("A2_dab", 10),
        ]:
            val = _to_float(ws4.cell(row_idx, col).value)
            if val is not None:
                result[key] = val
        return result

    cost_lines: list[dict[str, Any]] = []
    for r in range(10, 35):
        name = ws4.cell(r, 4).value
        if not name:
            continue
        label = str(name).strip()
        if label in ("RAZEM", "SUMA"):
            continue
        costs = row_costs(r)
        if not costs:
            continue
        section = "production"
        if r >= 24 and r <= 28:
            section = "print"
        elif r >= 29 and r <= 33:
            section = "packaging"
        elif r >= 34:
            section = "shipping"
        cost_lines.append({"name": label, "section": section, "costs": costs})

    def dump(name: str, payload: Any) -> None:
        (out / name).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    dump("materials.json", materials)
    dump("helpers.json", helpers)
    dump("price_table.json", price_table)
    dump("wood_defaults.json", wood_defaults)
    dump("settings.json", settings)
    dump("cost_lines.json", cost_lines)

    wb.close()
    return {
        "materials": len(materials),
        "price_table": len(price_table),
        "cost_lines": len(cost_lines),
    }


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="Import kalkulatora kosztów z .xlsm")
    parser.add_argument("xlsm", nargs="?", default=r"c:\Users\Skarabeusz\Downloads\do analizy.xlsm")
    args = parser.parse_args()
    stats = import_from_xlsm(args.xlsm)
    print("Zaimportowano:", stats)


if __name__ == "__main__":
    main()
