"""Import XLSX → JSON (legacy dev/maintenance). Runtime nie wymaga openpyxl."""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from .data_loader import data_dir, slugify

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]
    _IMPORT_ERR = exc
else:
    _IMPORT_ERR = None

_CATEGORY_BY_NAME: dict[str, str] = {
    "GUI / UI Premium": "UI",
    "Motion Director": "Motion",
    "Cursor Prompt Architect": "Cursor",
    "Code-aware Reviewer": "Cursor",
    "Shopify Snapshot Reviewer": "Shopify",
    "GicleeApp Architect": "GicleeApp",
    "Performance / Debug": "Performance",
    "Writer / Copy Premium": "Copy",
    "Veo / Flow / Image Prompt": "AI Prompt",
    "Medyczny ostrożny": "Medical",
}

_SHORT_LABEL_BY_NAME: dict[str, str] = {
    "GUI / UI Premium": "GUI Premium",
    "Motion Director": "Motion Director",
    "Cursor Prompt Architect": "Cursor Architect",
    "Code-aware Reviewer": "Code-aware Reviewer",
    "Shopify Snapshot Reviewer": "Shopify Snapshot Reviewer",
    "GicleeApp Architect": "GicleeApp Architect",
    "Performance / Debug": "Performance",
    "Writer / Copy Premium": "Copy Premium",
    "Veo / Flow / Image Prompt": "Veo Premium",
    "Medyczny ostrożny": "Medyczny ostrożny",
}

_EXTRA_ALIASES: dict[str, list[str]] = {
    "GUI / UI Premium": ["GUI Premium", "GUI / UI Premium"],
    "Cursor Prompt Architect": ["Cursor Architect", "Cursor Prompt Architect"],
    "Performance / Debug": ["Performance", "Performance / Debug"],
    "Writer / Copy Premium": ["Copy Premium", "Writer / Copy Premium"],
    "Veo / Flow / Image Prompt": ["Veo Premium", "Veo / Flow / Image Prompt"],
}


@dataclass(frozen=True)
class ParsedLegacyCatalog:
    source: str
    modes: tuple[dict[str, Any], ...]
    combinations: tuple[dict[str, Any], ...]


def _norm(value: Any) -> str:
    return str(value or "").strip()


def _default_source_path() -> Path:
    return data_dir() / "source" / "tryby_pracy_chatgpt_giclee_art.xlsx"


def _aliases_for_mode(name: str) -> list[str]:
    extras = list(_EXTRA_ALIASES.get(name, []))
    short = _SHORT_LABEL_BY_NAME.get(name, name)
    out: list[str] = []
    for candidate in [short, name, *extras]:
        c = _norm(candidate)
        if c and c not in out:
            out.append(c)
    return out


def _parse_modes_sheet(ws: Any) -> list[dict[str, Any]]:
    modes: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        number = row[0]
        name = _norm(row[1])
        if not name or number is None:
            continue
        purpose = _norm(row[2])
        focus = _norm(row[3])
        when_to_use = _norm(row[4])
        sample_command = _norm(row[5])
        simplest = _norm(row[6])
        aliases = _aliases_for_mode(name)
        short = aliases[0] if aliases else name
        modes.append(
            {
                "id": slugify(name),
                "order": int(number),
                "family": "legacy",
                "selectable": True,
                "name": name,
                "short_label": short,
                "aliases": aliases,
                "category": _CATEGORY_BY_NAME.get(name, "Inne"),
                "source_file": "",
                "purpose": purpose,
                "focus": focus,
                "when_to_use": when_to_use,
                "activation_profiles": [
                    {"id": "default", "label": "Domyślna", "command": sample_command.split("\n")[0] or short}
                ],
                "requires": [],
                "related_mode_ids": [],
                "distinction_note": simplest,
            }
        )
    return modes


def _alias_lookup(modes: list[dict[str, Any]]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for mode in modes:
        mode_id = mode["id"]
        keys = {mode["name"], *mode.get("aliases", [])}
        for key in keys:
            k = _norm(key)
            if k:
                lookup[k.casefold()] = mode_id
    return lookup


def _resolve_mode_ids(part: str, lookup: dict[str, str]) -> str | None:
    key = _norm(part).casefold()
    if key in lookup:
        return lookup[key]
    for prefix in ("tryb ",):
        if key.startswith(prefix):
            trimmed = key[len(prefix) :].strip()
            if trimmed in lookup:
                return lookup[trimmed]
    return None


def _split_combination_name(name: str) -> list[str]:
    return [_norm(p) for p in name.split(" + ") if _norm(p)]


def _parse_combinations_sheet(ws: Any, modes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    lookup = _alias_lookup(modes)
    combinations: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        name = _norm(row[0])
        if not name or name.casefold() == "kombinacja":
            continue
        best_for = _norm(row[1])
        delivers = _norm(row[2])
        usage_example = _norm(row[3])
        note = _norm(row[4])
        parts = _split_combination_name(name)
        mode_ids: list[str] = []
        missing: list[str] = []
        for part in parts:
            resolved = _resolve_mode_ids(part, lookup)
            if resolved:
                if resolved not in mode_ids:
                    mode_ids.append(resolved)
            else:
                missing.append(part)
        if missing:
            raise ValueError(
                f"Nie rozpoznano trybów w kombinacji «{name}»: {', '.join(missing)}"
            )
        combinations.append(
            {
                "id": slugify(name)[:80],
                "name": name,
                "mode_ids": mode_ids,
                "best_for": best_for,
                "delivers": delivers,
                "usage_example": usage_example,
                "note": note,
            }
        )
    return combinations


def parse_xlsx(source: str | Path | None = None) -> ParsedLegacyCatalog:
    if openpyxl is None:
        raise RuntimeError(
            "Brak biblioteki openpyxl — zainstaluj: pip install openpyxl"
        ) from _IMPORT_ERR

    path = Path(source) if source else _default_source_path()
    if not path.is_file():
        raise FileNotFoundError(f"Nie znaleziono pliku: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    if "Tryby pracy" not in wb.sheetnames:
        raise ValueError("Brak arkusza «Tryby pracy»")
    if "Kombinacje" not in wb.sheetnames:
        raise ValueError("Brak arkusza «Kombinacje»")

    modes = _parse_modes_sheet(wb["Tryby pracy"])
    combinations = _parse_combinations_sheet(wb["Kombinacje"], modes)
    return ParsedLegacyCatalog(
        source=path.name,
        modes=tuple(modes),
        combinations=tuple(combinations),
    )


def write_legacy_catalog(parsed: ParsedLegacyCatalog, output_dir: str | Path) -> dict[str, int]:
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    modes_payload = {
        "schema_version": 1,
        "source": parsed.source,
        "note": "Legacy import z XLSX — nie używać jako runtime v38",
        "modes": list(parsed.modes),
    }
    combos_payload = {
        "schema_version": 1,
        "source": parsed.source,
        "note": "Legacy import z XLSX — nie używać jako runtime v38",
        "combinations": list(parsed.combinations),
    }

    modes_path = out / "work_modes.json"
    combos_path = out / "combinations.json"
    modes_path.write_text(
        json.dumps(modes_payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    combos_path.write_text(
        json.dumps(combos_payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return {"modes": len(parsed.modes), "combinations": len(parsed.combinations)}


def import_from_xlsx(
    source: str | Path | None = None,
    *,
    output_dir: str | Path | None = None,
) -> dict[str, int]:
    """Parse-only domyślnie; zapis tylko po jawnym output_dir."""
    parsed = parse_xlsx(source)
    if output_dir is None:
        return {"modes": len(parsed.modes), "combinations": len(parsed.combinations)}
    return write_legacy_catalog(parsed, output_dir)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Legacy import XLSX → JSON (dev)")
    parser.add_argument("source", nargs="?", help="Ścieżka do pliku XLSX")
    parser.add_argument(
        "--output-dir",
        dest="output_dir",
        help="Katalog docelowy zapisu (wymagany do zapisu plików)",
    )
    args = parser.parse_args(argv)

    if args.output_dir:
        stats = import_from_xlsx(args.source, output_dir=args.output_dir)
        print(
            f"Zapisano legacy: {stats['modes']} trybów, "
            f"{stats['combinations']} kombinacji -> {args.output_dir}"
        )
    else:
        stats = import_from_xlsx(args.source)
        print(
            f"Sparsowano (bez zapisu): {stats['modes']} trybów, "
            f"{stats['combinations']} kombinacji"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
