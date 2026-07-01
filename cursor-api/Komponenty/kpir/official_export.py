"""Eksport urzędowy PKPiR — 19 kolumn (załącznik Dz.U. 2025 poz. 1299)."""

from __future__ import annotations

import csv
from pathlib import Path

from .annual_income import annual_income_breakdown
from .constants import OFFICIAL_COLUMN_HEADERS
from .entry_service import filter_entries
from .official_columns import (
    entry_to_official_row,
    monthly_cumulative_rows,
    sum_official_columns,
)
from .storage import documents_dir_for, load_settings


def _rows_for_period(year: int, month: int | None = None) -> list[dict]:
    settings = load_settings()
    entries = [
        e for e in filter_entries(year=year, month=month)
        if e.status in ("posted", "corrected")
    ]
    entries.sort(key=lambda x: (x.event_date, x.entry_number))
    return [entry_to_official_row(i, e, settings=settings) for i, e in enumerate(entries, 1)]


def export_official_pkpir_csv(year: int, month: int | None = None) -> Path:
    rows = _rows_for_period(year, month)
    out_dir = documents_dir_for("kpir", year, month or 12)
    suffix = f"{year:04d}_{month:02d}" if month else f"{year:04d}_roczny"
    out = out_dir / f"pkpir_urzedowy_{suffix}.csv"
    fieldnames = [k for k, _ in OFFICIAL_COLUMN_HEADERS]
    headers = {k: label for k, label in OFFICIAL_COLUMN_HEADERS}
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writerow(headers)
        for row in rows:
            w.writerow(row)
        if rows:
            totals = sum_official_columns(rows)
            w.writerow({k: totals.get(k, "") if k in totals else "SUMA" for k in fieldnames})
    return out


def export_official_pkpir_xlsx(year: int, month: int | None = None) -> Path:
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError as e:
        raise ImportError("Zainstaluj openpyxl: pip install openpyxl") from e

    settings = load_settings()
    rows = _rows_for_period(year, month)
    out_dir = documents_dir_for("kpir", year, month or 12)
    suffix = f"{year:04d}_{month:02d}" if month else f"{year:04d}_roczny"
    out = out_dir / f"pkpir_urzedowy_{suffix}.xlsx"
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "PKPiR"
    ws.append([label for _, label in OFFICIAL_COLUMN_HEADERS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(k) for k, _ in OFFICIAL_COLUMN_HEADERS])

    if month and settings.cumulative_monthly_sums:
        ws2 = wb.create_sheet("Narastająco")
        cum = monthly_cumulative_rows(rows, through_month=month)
        ws2.append(["Pozycja", "Kwota PLN"])
        for k, v in cum.items():
            ws2.append([k, v])

    if not month:
        ws3 = wb.create_sheet("Dochód roczny")
        inc = annual_income_breakdown(year, settings)
        for key in (
            "revenue", "inventory_opening", "purchase_goods", "purchase_side",
            "inventory_closing", "other_expenses_total", "total_costs", "income", "formula",
        ):
            ws3.append([key, inc.get(key)])

    wb.save(out)
    return out


def export_official_pkpir_pdf(year: int, month: int | None = None) -> Path:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as e:
        raise ImportError("Zainstaluj reportlab: pip install reportlab") from e

    settings = load_settings()
    rows = _rows_for_period(year, month)
    out_dir = documents_dir_for("kpir", year, month or 12)
    suffix = f"{year:04d}_{month:02d}" if month else f"{year:04d}_roczny"
    out = out_dir / f"pkpir_urzedowy_{suffix}.pdf"

    doc = SimpleDocTemplate(str(out), pagesize=landscape(A4), leftMargin=10 * mm, rightMargin=10 * mm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("PODATKOWA KSIĘGA PRZYCHODÓW I ROZCHODÓW", styles["Title"]),
        Paragraph(settings.seller_name or "", styles["Normal"]),
        Paragraph(settings.seller_address or "", styles["Normal"]),
        Paragraph(settings.activity_description or "", styles["Normal"]),
        Spacer(1, 8),
        Paragraph(f"Okres: {year}" + (f"-{month:02d}" if month else " (roczny)"), styles["Normal"]),
        Spacer(1, 8),
    ]

    headers = [label for _, label in OFFICIAL_COLUMN_HEADERS[:10]]
    data = [headers]
    for row in rows:
        data.append([
            str(row.get("lp", "")),
            row.get("event_date", ""),
            (row.get("document_number") or "")[:12],
            (row.get("contractor") or "")[:14],
            (row.get("description") or "")[:18],
            f"{row.get('revenue_goods', 0):.2f}",
            f"{row.get('revenue_other', 0):.2f}",
            f"{row.get('total_revenue', 0):.2f}",
            f"{row.get('purchase_goods', 0):.2f}",
            f"{row.get('other_expenses', 0):.2f}",
        ])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
    ]))
    story.append(table)
    doc.build(story)
    return out
