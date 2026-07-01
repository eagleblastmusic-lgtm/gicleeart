"""Eksporty KPiR — CSV, XLSX, PDF, dla księgowego."""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Any

from .constants import KPIR_COLUMN_LABELS
from .entry_service import filter_entries
from .storage import documents_dir_for, list_costs, list_entries
from .summary_service import monthly_summary, yearly_summary

_COMPONENT_DIR = Path(__file__).resolve().parent

KPIR_EXPORT_COLUMNS = [
    ("lp", "Lp."),
    ("event_date", "Data zdarzenia"),
    ("document_number", "Numer dowodu"),
    ("contractor", "Kontrahent"),
    ("contractor_address", "Adres kontrahenta"),
    ("description", "Opis"),
    ("revenue_goods", KPIR_COLUMN_LABELS["revenue_goods"]),
    ("revenue_other", KPIR_COLUMN_LABELS["revenue_other"]),
    ("total_revenue", "Razem przychód"),
    ("purchase_goods", KPIR_COLUMN_LABELS["purchase_goods"]),
    ("purchase_side", KPIR_COLUMN_LABELS["purchase_side"]),
    ("wages", KPIR_COLUMN_LABELS["wages"]),
    ("other_expenses", KPIR_COLUMN_LABELS["other_expenses"]),
    ("total_costs", "Razem koszty"),
    ("notes", "Uwagi"),
    ("source", "Źródło"),
    ("status", "Status"),
    ("original_currency", "Waluta"),
    ("original_amount", "Kwota oryg."),
    ("nbp_rate", "Kurs NBP"),
    ("nbp_rate_date", "Data kursu"),
    ("nbp_table_number", "Tabela NBP"),
    ("amount_pln", "Kwota PLN"),
    ("country", "Kraj"),
]


def _entry_row(lp: int, entry) -> dict[str, Any]:
    return {
        "lp": lp,
        "event_date": entry.event_date,
        "document_number": entry.document_number,
        "contractor": entry.contractor,
        "contractor_address": entry.contractor_address,
        "description": entry.description,
        "revenue_goods": entry.revenue_goods,
        "revenue_other": entry.revenue_other,
        "total_revenue": entry.total_revenue,
        "purchase_goods": entry.purchase_goods,
        "purchase_side": entry.purchase_side,
        "wages": entry.wages,
        "other_expenses": entry.other_expenses,
        "total_costs": entry.total_costs,
        "notes": entry.notes,
        "source": entry.source,
        "status": entry.status,
        "original_currency": entry.original_currency,
        "original_amount": entry.original_amount,
        "nbp_rate": entry.nbp_rate,
        "nbp_rate_date": entry.nbp_rate_date,
        "nbp_table_number": entry.nbp_table_number,
        "amount_pln": entry.amount_pln,
        "country": entry.country,
    }


def export_kpir_csv(year: int, month: int | None = None) -> Path:
    entries = filter_entries(year=year, month=month, status="posted")
    entries += [e for e in filter_entries(year=year, month=month) if e.status == "corrected"]
    out_dir = documents_dir_for("exports", year, month or 1)
    suffix = f"{year:04d}_{month:02d}" if month else f"{year:04d}"
    out = out_dir / f"kpir_{suffix}.csv"
    fieldnames = [c[0] for c in KPIR_EXPORT_COLUMNS]
    headers = {c[0]: c[1] for c in KPIR_EXPORT_COLUMNS}
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writerow(headers)
        for i, e in enumerate(sorted(entries, key=lambda x: x.event_date), 1):
            w.writerow(_entry_row(i, e))
    return out


def export_accountant_csv(year: int, month: int | None = None) -> Path:
    out_dir = documents_dir_for("exports", year, month or 1)
    suffix = f"{year:04d}_{month:02d}" if month else f"{year:04d}"
    out = out_dir / f"accountant_{suffix}.csv"
    rows: list[dict[str, Any]] = []
    for e in filter_entries(year=year, month=month):
        if e.status not in ("posted", "corrected"):
            continue
        rows.append({
            "type": e.entry_type,
            "date": e.event_date,
            "document": e.document_number,
            "shopify_order": e.shopify_order_name,
            "invoice_id": e.invoice_id,
            "amount_pln": e.amount_pln,
            "currency": e.original_currency,
            "nbp_rate": e.nbp_rate,
            "nbp_date": e.nbp_rate_date,
            "source": e.source,
            "category": e.category,
            "correction_of": e.linked_entry_id,
        })
    for c in list_costs():
        rows.append({
            "type": "cost_raw",
            "date": c.event_date,
            "document": c.document_number,
            "seller": c.seller,
            "amount_pln": c.amount_pln,
            "currency": c.currency,
            "kpir_status": c.kpir_status,
            "category": c.category,
        })
    if not rows:
        rows.append({"type": "empty", "note": "brak danych"})
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    return out


def export_kpir_xlsx(year: int, month: int | None = None) -> Path:
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError as e:
        raise ImportError("Zainstaluj openpyxl: pip install openpyxl") from e

    out_dir = documents_dir_for("exports", year, month or 1)
    suffix = f"{year:04d}_{month:02d}" if month else f"{year:04d}"
    out = out_dir / f"kpir_{suffix}.xlsx"
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "KPiR"
    headers = [c[1] for c in KPIR_EXPORT_COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    entries = filter_entries(year=year, month=month)
    for i, e in enumerate(sorted(
        [x for x in entries if x.status in ("posted", "corrected")],
        key=lambda x: x.event_date,
    ), 1):
        row = _entry_row(i, e)
        ws.append([row[c[0]] for c in KPIR_EXPORT_COLUMNS])

    if month:
        ws2 = wb.create_sheet("Podsumowanie miesiąca")
        summary = monthly_summary(year, month)
        ws2.append(["Pozycja", "Kwota PLN"])
        ws2.append(["Przychody razem", summary["revenue_total"]])
        ws2.append(["Koszty razem", summary["costs_total"]])
        ws2.append(["Dochód / strata", summary["income"]])
    else:
        ws2 = wb.create_sheet("Podsumowanie roku")
        summary = yearly_summary(year)
        ws2.append(["Miesiąc", "Przychody", "Koszty", "Dochód"])
        for m, data in summary["by_month"].items():
            ws2.append([m, data["revenue"], data["costs"], data["income"]])

    ws3 = wb.create_sheet("Koszty")
    ws3.append(["Data", "Dokument", "Sprzedawca", "Kategoria", "PLN", "Status KPiR"])
    for c in list_costs():
        ws3.append([c.event_date, c.document_number, c.seller, c.category, c.amount_pln, c.kpir_status])

    ws4 = wb.create_sheet("JPK_PKPIR")
    try:
        from .jpk_export import jpk_metadata
        meta = jpk_metadata(year, month or 1)
        ws4.append(["Format", meta["format"]])
        ws4.append(["Liczba wierszy", meta["row_count"]])
        ws4.append(["Uwaga", "Pełny XML: eksport JPK_PKPIR XML"])
    except Exception:
        ws4.append(["Uwaga", "JPK_PKPIR — eksport XML osobno"])

    wb.save(out)
    return out


def export_accountant_package(year: int, month: int) -> Path:
    """Pakiet dla księgowego: CSV + XLSX + PDF + JPK XML w jednym folderze."""
    import shutil

    base = documents_dir_for("exports", year, month)
    pkg = base / f"pakiet_ksiegowy_{year:04d}_{month:02d}"
    pkg.mkdir(parents=True, exist_ok=True)

    paths = [
        export_kpir_csv(year, month),
        export_accountant_csv(year, month),
    ]
    try:
        paths.append(export_kpir_xlsx(year, month))
    except ImportError:
        pass
    try:
        paths.append(export_kpir_pdf(year, month))
    except ImportError:
        pass
    try:
        from .jpk_export import export_jpk_pkpir_xml
        paths.append(export_jpk_pkpir_xml(year, month))
    except Exception:
        pass

    for src in paths:
        if src.is_file():
            shutil.copy2(src, pkg / src.name)

    readme = pkg / "README.txt"
    readme.write_text(
        f"Pakiet księgowy KPiR — {year}-{month:02d}\n"
        f"Wygenerowano: {datetime.now().isoformat(timespec='seconds')}\n"
        f"Pliki: {', '.join(p.name for p in pkg.iterdir() if p.is_file())}\n",
        encoding="utf-8",
    )
    return pkg


def export_kpir_pdf(year: int, month: int) -> Path:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as e:
        raise ImportError("Zainstaluj reportlab: pip install reportlab") from e

    def _esc(text: str) -> str:
        return (
            str(text or "")
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )

    out_dir = documents_dir_for("kpir", year, month)
    out = out_dir / f"kpir_{year:04d}_{month:02d}.pdf"
    summary = monthly_summary(year, month)
    entries = filter_entries(year=year, month=month, status="posted")

    margin = 18 * mm
    page_w, _page_h = A4
    content_w = page_w - 2 * margin

    doc = SimpleDocTemplate(
        str(out),
        pagesize=A4,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin,
    )
    styles = getSampleStyleSheet()
    cell_left = ParagraphStyle(
        "kpir_cell_l", parent=styles["Normal"], fontSize=8, leading=10, wordWrap="CJK",
    )
    cell_right = ParagraphStyle(
        "kpir_cell_r", parent=cell_left, alignment=2,
    )
    header_style = ParagraphStyle(
        "kpir_hdr", parent=styles["Normal"], fontSize=8, leading=10, textColor=colors.whitesmoke,
    )

    story = [
        Paragraph(f"KPiR — {year}-{month:02d}", styles["Title"]),
        Spacer(1, 12),
        Paragraph(
            f"Przychody: {summary['revenue_total']:.2f} PLN | "
            f"Koszty: {summary['costs_total']:.2f} PLN | "
            f"Dochód: {summary['income']:.2f} PLN",
            styles["Normal"],
        ),
        Spacer(1, 12),
    ]

    col_widths = [
        content_w * 0.11,  # Data
        content_w * 0.30,  # Dokument — długie numery CYK/REC-…
        content_w * 0.27,  # Kontrahent
        content_w * 0.16,  # Przychód
        content_w * 0.16,  # Koszty
    ]

    table_data: list[list] = [[
        Paragraph(_esc("Data"), header_style),
        Paragraph(_esc("Dokument"), header_style),
        Paragraph(_esc("Kontrahent"), header_style),
        Paragraph(_esc("Przychód"), header_style),
        Paragraph(_esc("Koszty"), header_style),
    ]]
    for e in sorted(entries, key=lambda x: x.event_date):
        table_data.append([
            Paragraph(_esc(e.event_date[:10]), cell_left),
            Paragraph(_esc(e.document_number), cell_left),
            Paragraph(_esc(e.contractor or ""), cell_left),
            Paragraph(_esc(f"{e.total_revenue:.2f}"), cell_right),
            Paragraph(_esc(f"{e.total_costs:.2f}"), cell_right),
        ])

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("ALIGN", (3, 1), (4, -1), "RIGHT"),
    ]))
    story.append(table)
    doc.build(story)
    return out
